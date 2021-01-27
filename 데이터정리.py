import os.path
import glob
import pandas as pd
from matplotlib import pyplot as plt
import openpyxl as xl
from openpyxl import load_workbook
from openpyxl import Workbook
import pyperclip as clp
import pyautogui as gui
import time
import csv

data=pd.read_csv("D:/EMG머신러닝연구/윤이나 학생 자료/EMG_data/brushedData2/EMGdata-작업용.csv")
wb=xl.load_workbook("D:/EMG머신러닝연구/윤이나 학생 자료/EMG_data/brushedData2/EMGdata1.xlsx")
ws=wb.active
results=[]
for row in ws.iter_rows(min_row=0,max_row=len(data),min_col=0,max_col=len(data.columns)):
    result=[]
    for cell in row:
        if cell.value!=0:
            result.append(cell.value)
    results.append(result)
# for r in ws.rows:
#     result=[]
#     if r.value!=0: #데이터가 비어있으면 단계를 건너뛰는 코드, 빈데이터를 none으로 처리
#         for c in r:
#             result.append(c.value)
#     results.append(result)
wb=Workbook()
ws=wb.active
for i in results:
    ws.append(i)
wb.save("results.xlsx")
# # data.loc[2,:]==0
# for i in range(len(data)):
#     for j in range(len(data.columns)):
#         if data.iloc[i,j:j+5]==0:


# if
# reqd_Index = df[df['Sales']>=300].index.tolist()
# print(reqd_Index)
# files2 = glob.glob('D:/EMG머신러닝연구/윤이나 학생 자료/EMG_data/brushedData2/EMGdata-작업용.csv')
# list = []
# for file in files2:
#     file = open(file, 'r', encoding='utf-8')
#     file_a = csv.reader(file)
#     for line in file_a:
#         print(line)

# wb = xl.load_workbook('D:/EMG머신러닝연구/윤이나 학생 자료/EMG_data/brushedData2/EMGdata1.xlsx')  # 엑셀데이터 불러오기
# ws = wb.active  # 엑셀데이터가 행단위로 리스트형태로 저장됨
# for r in ws.rows: #모든셀의 데이터를 하나씩 불러와서 명령문 실행



